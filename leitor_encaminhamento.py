import pdfplumber
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

# Caminho do PDF
caminho_pdf = "./import_pdf/encaminhamentos.pdf"

# Lista para armazenar os dados de cada aluno
dados_alunos = []

# Define a ordem desejada das colunas
colunas_ordem = [
    "Matrícula", "Aluno", "Escola", "Etapa", "Nível/Fase", "Turma", "Professor",
    "Nome da Mãe", "Nome do Pai", "Responsável", "Parentesco", "Endereço", "Bairro",
    "Telefone", "Motivo do Encaminhamento", "Psicólogo(a)", "Fonoaudiólogo(a)",
    "Psicopedagogo(a)", "Serviço Social", "Status"
]

with pdfplumber.open(caminho_pdf) as pdf:
    for pagina in pdf.pages:
        texto = pagina.extract_text()
        if not texto or "1. IDENTIFICAÇÃO" not in texto:
            continue
        texto_aluno = texto.split("1. IDENTIFICAÇÃO", 1)[1]

        if texto:
            # Inicializa o dicionário para armazenar os dados do aluno
            dados = {}

            # Expressões regulares para extrair as informações
            escola_info = re.search(r'ESCOLA:\s*\d+\s*-\s*(.*?)\n', texto)
            aluno =  re.search(r'ALUNO:\s*(\d+)\s*-\s*(\d+)\s+(.*)', texto_aluno)
            filiacao = re.search(r'FILIAÇÃO:\s*(.*?)\n(?:([^\n]*)\n)?', texto_aluno)
            responsavel_match = re.search(r'RESPONSÁVEL:\s*(.*?)\n', texto_aluno)
            parentesco_match = re.search(r'GRAU PARENTESCO:\s*(.*?)\n', texto_aluno)
            endereco_match = re.search(r'ENDEREÇO:\s*(.*?)\n', texto_aluno)
            bairro_match = re.search(r'BAIRRO:\s*(.*?)\n', texto_aluno)
            fone_match = re.search(r'FONE:\s*(.*?)\n', texto_aluno)
            serie_completa = re.search(r'SÉRIE:\s*(.*?)\n', texto_aluno)
            turma_match = re.search(r'TURMA:\s*(\w+)', texto_aluno)
            motivo_match = re.search(r'MOTIVO DO ENCAMINHAMENTO\s*\n\s*(.*?)(?=\n4\.\s+AÇÕES DESENVOLVIDAS)', texto_aluno, re.DOTALL)
            professor = re.search(r'PROFESSOR:\s*(.*?)\n', texto_aluno)

            # Adiciona os dados extraídos ao dicionário
            if escola_info:
                dados['Escola'] = escola_info.group(1).strip()
            if aluno:
                cod_num = aluno.group(1)
                cod_dig = aluno.group(2)
                nome_final = aluno.group(3).strip()
                codigo_final = f"{cod_num}{cod_dig}"
            else:
                codigo_final = ""
                nome_final = ""
            dados['Matrícula'] = codigo_final
            dados['Aluno'] = nome_final
            if professor:
                dados['Professor'] = professor.group(1).strip()
            if serie_completa:
                serie_texto = serie_completa.group(1).strip()
                nivel_match = re.search(r'(\d+\s+\w+)$', serie_texto)
                if nivel_match:
                    dados['Etapa'] = re.sub(r'\s*-\s*' + re.escape(nivel_match.group(1)) + r'$', '', serie_texto).strip()
                    dados['Nível/Fase'] = nivel_match.group(1).strip()
                else:
                    dados['Etapa'] = serie_texto
                    dados['Nível/Fase'] = ''
            if filiacao:
                nome_mae = filiacao.group(1).strip()
                nome_pai = filiacao.group(2).strip() if filiacao.group(2) and not filiacao.group(2).strip().startswith("RESPONSÁVEL") else ""
                dados['Nome da Mãe'] = nome_mae
                dados['Nome do Pai'] = nome_pai
            else:
                dados['Nome da Mãe'] = ""
                dados['Nome do Pai'] = ""
            if responsavel_match:
                resp = responsavel_match.group(1).strip()
                dados['Responsável'] = "" if resp.startswith("GRAU PARENTESCO") else resp
            if parentesco_match:
                parentesco = parentesco_match.group(1).strip()
                dados['Parentesco'] = "" if parentesco.startswith("ENDEREÇO") else parentesco
            if endereco_match:
                dados['Endereço'] = endereco_match.group(1).strip()
            if bairro_match:
                dados['Bairro'] = bairro_match.group(1).strip()
            if fone_match:
                dados['Telefone'] = fone_match.group(1).strip()
            if turma_match:
                dados['Turma'] = turma_match.group(1).strip()
            if motivo_match:
                dados['Motivo do Encaminhamento'] = motivo_match.group(1).strip()

            dados["Psicólogo(a)"] = ""
            dados["Fonoaudiólogo(a)"] = ""
            dados["Psicopedagogo(a)"] = ""
            dados["Serviço Social"] = ""
            dados["Status"] = ""


            for coluna in colunas_ordem:
                if coluna not in dados:
                    dados[coluna] = ""
            
            # Adiciona o dicionário à lista se contiver dados
            if dados:
                dados_alunos.append(dados)

# Cria o DataFrame com colunas na ordem definida
df = pd.DataFrame(dados_alunos)[colunas_ordem]

# Salva os dados em um arquivo Excel
df.to_excel("./export/dados_encaminhamentos_alunos.xlsx", index=False)


arquivo_excel = "./export/dados_encaminhamentos_alunos.xlsx"
wb = load_workbook(arquivo_excel)
ws = wb.active

# Carrega os nomes do arquivo assessores.xlsx
df_assessores = pd.read_excel("./assessores/assessores.xlsx")
nomes_assessores = df_assessores["Nomes"].dropna().tolist()

aba_assessores = wb.create_sheet("Assessor_Nomes")
for i, nome in enumerate(nomes_assessores, start=1):
    aba_assessores[f"A{i}"] = nome

aba_assessores.sheet_state = "hidden"

intervalo_assessores = f"Assessor_Nomes!$A$1:$A${len(nomes_assessores)}"
dv_assessor = DataValidation(type="list", formula1=f"={intervalo_assessores}", allow_blank=True)
dv_assessor.error = "Escolha um nome válido"
dv_assessor.prompt = "Selecione o assessor da lista"


status_options = ["RECEBIDOS", "ATENDIDOS", "CONCLUÍDOS", "RELATÓRIO FINALIZADO"]
dv_status = DataValidation(
    type="list",
    formula1=f'"{",".join(status_options)}"',
    allow_blank=True
)
dv_status.error = "Escolha um valor válido"
dv_status.prompt = "Selecione um status da lista"


linha_inicial = 2
linha_final = ws.max_row
coluna_psicologo = "P"
coluna_fonodiologo = "Q"
coluna_psicopedagogo = "R"
coluna_servicosocial = "S"
coluna_status = "T"

# Preenche automaticamente com "RECEBIDOS"
for row in range(linha_inicial, linha_final + 1):
    ws[f"{coluna_status}{row}"] = "RECEBIDOS"

ws.add_data_validation(dv_status)
dv_status.ranges.add(f"{coluna_status}{linha_inicial}:{coluna_status}{linha_final}")

ws.add_data_validation(dv_assessor)
dv_assessor.ranges.add(f"{coluna_psicologo}{linha_inicial}:{coluna_psicologo}{linha_final}")
dv_assessor.ranges.add(f"{coluna_fonodiologo}{linha_inicial}:{coluna_fonodiologo}{linha_final}")
dv_assessor.ranges.add(f"{coluna_psicopedagogo}{linha_inicial}:{coluna_psicopedagogo}{linha_final}")
dv_assessor.ranges.add(f"{coluna_servicosocial}{linha_inicial}:{coluna_servicosocial}{linha_final}")

wb.save(arquivo_excel)

print("Arquivo 'dados_encaminhamentos_alunos.xlsx' criado com sucesso!")
