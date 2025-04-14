import pdfplumber
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

# Caminho do PDF
caminho_pdf = "./import_pdf/ficai.pdf"

# Lista para armazenar os dados de cada aluno
dados_alunos = []

# Define a ordem desejada das colunas
colunas_ordem = [
    "Matrícula", "Aluno", "Escola", "Etapa", "Nível/Fase", "Turma", "Turno",
    "Professor", "Nome do Pai", "Nome da Mãe", "Endereço Residencial",
    "Ponto de Referência", "Telefones", 
    "Assessor Responsável Teleresgate", "Assessor Responsável Visita",
    "Status"
]

with pdfplumber.open(caminho_pdf) as pdf:
    for pagina in pdf.pages:
        texto = pagina.extract_text()
        if texto:
            # Inicializa o dicionário para armazenar os dados do aluno
            dados = {}
            linhas = texto.split("\n")

            nome_pai = ""
            nome_mae = ""
            for i, linha in enumerate(linhas):
                if "NOME DO PAI" in linha and i > 0:
                    anterior = linhas[i - 1].strip()
                    if anterior and "NOME" not in anterior:
                        nome_pai = anterior
                if "NOME DA MÃE" in linha and i > 0:
                    anterior = linhas[i - 1].strip()
                    if anterior and "ENDEREÇO" not in anterior:
                        nome_mae = anterior

            # Expressões regulares para extrair as informações
            escola_info = re.search(r'NOME\s+\d+\s+-\s+(.*?)\s+-\s+POLO/DDZ:\s+(.*?)\n', texto)
            aluno_match = re.findall(r'NOME\s+(\d+)\s*-\s*(\d+)\s*-\s+(.*?)\n', texto)
            serie_completa = re.search(r'SÉRIE\s+(.*?)\n', texto)
            turma = re.search(r'TURMA\s+([A-Z])', texto)
            turno = re.search(r'TURNO\s+(\w+)', texto)
            professor = re.search(r'NOME DO PROFESSOR\s+(.*?)\n', texto)
            pai = re.search(r'NOME DO PAI\s+(.+)', texto)
            mae = re.search(r'NOME DA MÃE\s+(.+)', texto)
            endereco = re.search(r'ENDEREÇO RESIDENCIAL\s+(.+)', texto)
            # Ponto de Referência
            referencia_match = re.search(r'PONTO DE REFERENCIA\s*(.*?)\n', texto)
            referencia = referencia_match.group(1).strip() if referencia_match else ""
            telefones = re.search(r'Fone:\s*(.+)', texto)

            # Adiciona os dados extraídos ao dicionário
            if escola_info:
                dados['Escola'] = escola_info.group(1).strip()
            if aluno_match:
                cod_num, cod_dig, nome_aluno = aluno_match[-1]
                codigo_final = f"{cod_num.strip()}{cod_dig.strip()}"
                nome_final = nome_aluno.strip()
            else:
                codigo_final = ""
                nome_final = ""
            dados['Matrícula'] = codigo_final
            dados['Aluno'] = nome_final
            if serie_completa:
                serie_texto = serie_completa.group(1).strip()
                nivel_match = re.search(r'(\d+\s+\w+)$', serie_texto)
                if nivel_match:
                    dados['Etapa'] = serie_texto.replace(nivel_match.group(1), '').strip()
                    dados['Nível/Fase'] = nivel_match.group(1).strip()
                else:
                    dados['Etapa'] = serie_texto
                    dados['Nível/Fase'] = ''
            if turma:
                dados['Turma'] = turma.group(1).strip()
            if turno:
                dados['Turno'] = turno.group(1).strip()
            if professor:
                dados['Professor'] = professor.group(1).strip()
            if pai:
                nome_pai = pai.group(1).strip()
                if nome_pai and not nome_pai.upper().startswith("NOME"):
                    dados["Nome do Pai"] = nome_pai
                else:
                    dados["Nome do Pai"] = ""
            else:
                dados['Nome do Pai'] = ""
            if mae:
                nome_mae = mae.group(1).strip()
                if nome_mae and not nome_mae.upper().startswith("ENDEREÇO"):
                    dados["Nome da Mãe"] = nome_mae
                else:
                    dados["Nome da Mãe"] = ""
            else:
                dados['Nome da Mãe'] = ""
            if endereco:
                dados['Endereço Residencial'] = endereco.group(1).strip()
            if referencia and not referencia.lower().startswith("fone"):
                dados["Ponto de Referência"] = referencia
            if telefones:
                fones_texto = telefones.group(1)
                numeros = re.findall(r'\b(?:\d{2}\s)?\d{4,5}\d{4}\b', fones_texto)
                dados["Telefones"] = ", ".join(numeros)

            dados["Assessor Responsável Teleresgate"] = ""
            dados["Assessor Responsável Visita"] = ""
            dados["Status"] = ""

            for coluna in colunas_ordem:
                if coluna not in dados:
                    dados[coluna] = ""

            # Adiciona o dicionário à lista se contiver dados
            if dados:
                dados_alunos.append(dados)

# Cria um DataFrame com os dados
df = pd.DataFrame(dados_alunos)[colunas_ordem]

# Salva os dados em um arquivo Excel
df.to_excel("./export/dados_ficai_alunos.xlsx", index=False)

arquivo_excel = "./export/dados_ficai_alunos.xlsx"
wb = load_workbook(arquivo_excel)
ws = wb.active

# Carrega os nomes do arquivo assessores.xlsx
df_assessores = pd.read_excel("./assessores/assessores.xlsx")
nomes_assessores = df_assessores["Nomes"].dropna().tolist()

aba_assessores = wb.create_sheet("Assessor_Nomes")
for i, nome in enumerate(nomes_assessores, start=1):
    aba_assessores[f"A{i}"] = nome

aba_assessores.sheet_state = "hidden"

intervalo_assessores = f"Assessor_Nomes!$A$1:$A${len(nomes_assessores)}"
dv_assessor = DataValidation(type="list", formula1=f"={intervalo_assessores}", allow_blank=True)
dv_assessor.error = "Escolha um nome válido"
dv_assessor.prompt = "Selecione o assessor da lista"

status_options = ["RECEBIDOS", "ATENDIDOS", "CONCLUÍDOS", "RELATÓRIO FINALIZADO"]
dv_status = DataValidation(
    type="list",
    formula1=f'"{",".join(status_options)}"',
    allow_blank=True
)
dv_status.error = "Escolha um valor válido"
dv_status.prompt = "Selecione um status da lista"


linha_inicial = 2
linha_final = ws.max_row
coluna_visita = "N"
coluna_teleresgate = "O"
coluna_status = "P"


# Preenche automaticamente com "RECEBIDOS"
for row in range(linha_inicial, linha_final + 1):
    ws[f"{coluna_status}{row}"] = "RECEBIDOS"

ws.add_data_validation(dv_status)
dv_status.ranges.add(f"{coluna_status}{linha_inicial}:{coluna_status}{linha_final}")

ws.add_data_validation(dv_assessor)
dv_assessor.ranges.add(f"{coluna_teleresgate}{linha_inicial}:{coluna_teleresgate}{linha_final}")
dv_assessor.ranges.add(f"{coluna_visita}{linha_inicial}:{coluna_visita}{linha_final}")

# Salva de novo o arquivo
wb.save(arquivo_excel)

print("Arquivo 'dados_ficai_alunos.xlsx' criado com sucesso!")
