import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

# Caminhos dos arquivos
caminho_dados = "./export/dados_ficai_alunos.xlsx"
caminho_abandono = "./potencial-abandono/potencial_abandono_matriculas.xlsx"

# Carrega os dados das planilhas
df_dados = pd.read_excel(caminho_dados)
df_abandono = pd.read_excel(caminho_abandono)


# Converte colunas de matrícula para string para garantir comparação correta
df_dados["Matrícula"] = df_dados["Matrícula"].astype(str)
df_abandono["Matrícula"] = df_abandono["Matrícula"].astype(str)

# Atualiza o status para 'ATENDIDOS' se a matrícula estiver na planilha de abandono
df_dados.loc[df_dados["Matrícula"].isin(df_abandono["Matrícula"]), "Status"] = "ATENDIDOS"


# Salva novamente a planilha atualizada
df_dados.to_excel(caminho_dados, index=False)

# --- Reaplica as validações com openpyxl ---
wb = load_workbook(caminho_dados)
ws = wb.active

# Validação de status
status_options = ["RECEBIDOS", "ATENDIDOS", "CONCLUÍDOS", "RELATÓRIO FINALIZADO"]
dv_status = DataValidation(
    type="list",
    formula1=f'"{",".join(status_options)}"',
    allow_blank=True
)
dv_status.error = "Escolha um valor válido"
dv_status.prompt = "Selecione um status da lista"

# Validação de assessores
df_assessores = pd.read_excel("./assessores/assessores.xlsx")
nomes_assessores = df_assessores["Nomes"].dropna().tolist()

aba_assessores = wb.create_sheet("Assessor_Nomes")
for i, nome in enumerate(nomes_assessores, start=1):
    aba_assessores[f"A{i}"] = nome
aba_assessores.sheet_state = "hidden"

intervalo_assessores = f"Assessor_Nomes!$A$1:$A${len(nomes_assessores)}"
dv_assessor = DataValidation(type="list", formula1=f"={intervalo_assessores}", allow_blank=True)
dv_assessor.error = "Escolha um nome válido"
dv_assessor.prompt = "Selecione o assessor da lista"

# Colunas
linha_inicial = 2
linha_final = ws.max_row
coluna_visita = "N"
coluna_teleresgate = "O"
coluna_status = "P"

# Reaplica as listas suspensas (sem alterar valores existentes)
ws.add_data_validation(dv_status)
dv_status.ranges.add(f"{coluna_status}{linha_inicial}:{coluna_status}{linha_final}")

ws.add_data_validation(dv_assessor)
dv_assessor.ranges.add(f"{coluna_teleresgate}{linha_inicial}:{coluna_teleresgate}{linha_final}")
dv_assessor.ranges.add(f"{coluna_visita}{linha_inicial}:{coluna_visita}{linha_final}")

# Salva final
wb.save(caminho_dados)

print("Status atualizado com sucesso na planilha 'dados_ficai_alunos.xlsx'.")